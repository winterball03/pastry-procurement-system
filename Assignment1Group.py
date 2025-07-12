import openpyxl

def display_menu():
    print('-'*60)
    print('            Welcome to Sunflower Bakery System!')
    print('-'*60)
    print('Please choose an option from below.')
    print('001 = Create bakery product')
    print('002 = List out the bakery product')
    print('003 = Edit bakery product')
    print('004 = Delete bakery product')
    print('005 = Order')
    print('006 = Print bill statement for customer')
    print('007 = Generate Report for each months')
    print('008 = Exit System')
    print(' ')
    print('='*60)

#Main program loop
while True:
    display_menu()

    option = input('Enter option code: ')

    #Option 1 - Create menu
    if option == '001':
        print('-'*60)
        print('Option 001 - Create Bakery Product')
        print('-'*60)

        #Function to add a new menu to the Excel workbook
        def add_menu_to_excel(menu_id, menu_name, menu_category, menu_desc, menu_price, menu_recommendation):

            #Load the existing workbook
            workbook = openpyxl.load_workbook('Bakery.xlsx')  

            #Select the existing worksheet
            worksheet = workbook['MasterList']


            #Find the next available row in the worksheet
            next_row = len(worksheet['A']) + 1

            #Write menu data to the worksheet
            worksheet[f'A{next_row}'] = menu_id
            worksheet[f'B{next_row}'] = menu_name
            worksheet[f'C{next_row}'] = menu_category
            worksheet[f'D{next_row}'] = menu_desc
            worksheet[f'E{next_row}'] = menu_price
            worksheet[f'F{next_row}'] = menu_recommendation

            #Save the workbook
            workbook.save('Bakery.xlsx')

        #Define valid categories and chef recommendations
        valid_categories = ['Bread', 'Dessert', 'Cakes']
        valid_recommendations = ['Yes', 'No']

        #Function to validate menu price as a float
        def validate_menu_price(price):
            try:
                float_price = float(price)
                if float_price < 0:
                    raise ValueError("Price cannot be negative.")
                return True, float_price
            except ValueError:
                return False, None

        #Function to get valid input for category
        def get_valid_category():
            while True:
                category = input('Please choose a category (Bread/Dessert/Cakes): ').strip().capitalize()
                if category in valid_categories:
                    return category
                else:
                    print('Invalid category. Please choose from Bread, Dessert, or Cakes.')

        #Function to get valid input for chef recommendation
        def get_valid_recommendation():
            while True:
                recommendation = input('Chef Recommendation (Yes/No): ').strip().capitalize()
                if recommendation in valid_recommendations:
                    return recommendation
                else:
                    print('Invalid input. Please choose Yes or No.')

        menu_id = input('Menu ID: ')
        menu_category = get_valid_category()
        menu_name = input('Menu Name: ')
        menu_desc = input('Menu Description: ')

        #Get valid menu price
        while True:
            menu_price = input('Menu Price: ')
            is_valid, float_price = validate_menu_price(menu_price)
            if is_valid:
                break
            else:
                print('Invalid price. Please enter a valid numeric price.')

        menu_recommend = get_valid_recommendation()

        #Save the new menu into Excel
        add_menu_to_excel(menu_id, menu_category, menu_name, menu_desc, float_price, menu_recommend)
        print(' ')
        print('='*60)
        print('New menu has been successfully added into system.')
        print('Returning to main menu...')
        print(' ')

        pass

    #Option 2 - List menu
    elif option == '002':
        print('-'*60)
        print('Option 002 - List of Bakery Products')
        print('-'*60)

        #Function to list bakery products from the Excel workbook
        def list_bakery_products():
            try:
                workbook = openpyxl.load_workbook('Bakery.xlsx', data_only=True)  #Open the workbook in data-only mode
                worksheet = workbook['MasterList']  #Select the worksheet

                #Initialize a counter for valid rows
                valid_rows = 0

                #Define the expected number of columns
                expected_columns = 6

                #Iterate through rows and print bakery product details starting from the third row
                for row_number, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):  #Start from the third row
                    try:
                        #Check if the row contains at least the expected number of values
                        if len(row) >= expected_columns and all(row[:expected_columns]):  
                            valid_rows += 1
                            (menu_id, menu_category, menu_name, menu_desc, menu_price, menu_recommendation) = row[:expected_columns]
                            print('{:<10} {:<20} {:<15} {:<40} {:<10} {:<20}'.format(menu_id, menu_category, menu_name, menu_desc, menu_price, menu_recommendation))
                        else:
                            print(f"Error in Row {row_number}: Row does not contain the expected number of values")
                            print(f"Row {row_number} Content: {row}")
                    except ValueError:
                        print(f"Error in Row {row_number}: Error while processing row")
                        print(f"Row {row_number} Content: {row}")

                #Check if no valid rows were found
                if valid_rows == 0:
                    print("No valid bakery products found in the worksheet.")
                    
            except FileNotFoundError:
                print("No items in menu, please add some menu items first.")
                
        list_bakery_products()
        print(' ')
        print('='*60)
        print('All bakery products have been listed.')
        print('Returning to main menu...')
        print(' ')

        pass

    #Option 3 - Edit menu
    elif option == '003':
        print('-'*60)
        print('Option 003 - Edit bakery product')
        print('-'*60)

        def edit_bakery_product(wb, ws):
            #Prompt the user for the menu ID they want to edit
            menu_id_to_edit = input("Enter the Menu ID to edit: ")

            #Define the expected number of columns
            expected_columns = 6

            #Search for the bakery product
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2): 
                (menu_id, menu_category, menu_name, menu_desc, menu_price, menu_recommend) = row[:expected_columns]
                if menu_id == menu_id_to_edit:
                    
                    #Print out the original product information
                    print("Bakery Product was found: ")
                    print("Product ID:", menu_id)
                    print("Category:", menu_category)
                    print("Name:", menu_name)
                    print("Description:", menu_desc)
                    print("Price:", menu_price)
                    print("Recommendation:", menu_recommend)
                    print(' ')

                    #Prompt the user to input new data
                    print('-'*60)
                    new_category = input("Please choose a new category (Bread/Dessert/Cakes): ")
                    new_name = input("Enter new product name: ")
                    new_desc = input("Enter new description: ")
                    new_price = float(input("Enter new price: "))
                    new_recommend = input("Enter new chef recommendation (Yes/No): ")

                    #Update the values in the row
                    updated_row = (menu_id, new_category, new_name, new_desc, new_price, new_recommend)
                    for col_number, value in enumerate(updated_row, start=1):
                        ws.cell(row=row_number, column=col_number, value=value)
                
                    print(' ')
                    print('='*60)
                    print("Bakery Product has been Updated!")
                    print('Returning to main menu...')
                    print(' ')
                    break  #Exit the loop once the product is found and updated

            else:
                print("Bakery Product with ID", product_id_to_edit, " was not found.")

            #Save the changes to the Excel file
            wb.save('Bakery.xlsx')

        #Open the Excel workbook and select the worksheet where your bakery products are stored
        wb = openpyxl.load_workbook('Bakery.xlsx')
        ws = wb['MasterList']

        #Call the edit_bakery_product function to start the editing process
        edit_bakery_product(wb, ws)

        #Close the Excel workbook
        wb.close()

        pass

    #Option 4 - Delete bakery product
    elif option == '004':
        print('-'*60)
        print('Option 003 - Delete bakery product')
        print('-'*60)

        def delete_bakery_product(wb, ws):
            #Prompt the user for the menu ID they want to delete
            menu_id_to_delete = input("Enter the Menu ID to delete: ")

            #Define the expected number of columns
            expected_columns = 6

            #Create a new worksheet to copy the data without the deleted row
            new_ws = wb.create_sheet(title='TempSheet')

            #Initialize a variable to keep track of the row number
            row_number = 1  #Start from the first row

            #Iterate through the rows to find and copy the matching bakery product
            found = False
            for row in ws.iter_rows(min_row=1, values_only=True):
                (menu_id, menu_category, menu_name, menu_desc, menu_price, menu_recommend) = row[:expected_columns]
                if menu_id == menu_id_to_delete:
                    found = True
                    continue  #Skip copying this row

                #Copy the row to the new worksheet
                new_ws.append(row)
                row_number += 1

            #Delete the original worksheet
            wb.remove(ws)

            #Rename the new worksheet to match the original
            new_ws.title = 'MasterList'

            if found:
                print(' ')
                print('=' * 60)
                print("Bakery Product with Menu ID", menu_id_to_delete, "has been deleted.")
                print('Returning to main menu...')
                print(' ')
            else:
                print("Bakery Product with Menu ID", menu_id_to_delete, "not found.")

        #Open the Excel workbook and select the worksheet where your bakery products are stored
        wb = openpyxl.load_workbook('Bakery.xlsx')
        ws = wb['MasterList']

        #Call the delete_bakery_product function to start the deletion process
        delete_bakery_product(wb, ws)

        #Save the changes to the Excel file
        wb.save('Bakery.xlsx')

        #Close the Excel workbook
        wb.close()

        pass

    #Option 5 - Create customer record
    elif option == '005':
        print('-'*60)
        print('Option 005 - Order')
        print('-'*60)

        #Function to validate date and time format
        from datetime import datetime
        
        def validate_date_time(date_time):
            try:
                datetime.strptime(date_time, '%Y-%m-%d %H:%M:%S')
                return True
            except ValueError:
                return False

        #Function to validate amount as a float
        def validate_amount(amount):
            try:
                float_amount = float(amount)
                if float_amount < 0:
                    raise ValueError("Amount cannot be negative.")
                return True, float_amount
            except ValueError:
                return False, None

        #Function to validate payment method
        def validate_payment_method(payment_method):
            valid_payment_methods = ['Credit Card', 'Debit Card', 'Cash', 'Online Payment']
            return payment_method in valid_payment_methods

        #Function to validate order delivered input
        def validate_order_delivered(order_delivered):
            return order_delivered.lower() in ['yes', 'no']

        #Function to get valid input for a field
        def get_valid_input(prompt, validator):
            while True:
                user_input = input(prompt)
                if validator(user_input):
                    return user_input
                else:
                    print('Invalid input. Please enter a valid value.')
                    
        #Open the Excel workbook
        workbook = openpyxl.load_workbook('Bakery.xlsx')

        #Select the worksheet
        worksheet = workbook['Customer_Detail']

        # Get input from the user
        cus_date_n_time = get_valid_input('Date and Time (YYYY-MM-DD HH:MM:SS): ', validate_date_time)
        order_id = input('Order ID: ')
        cus_name = input('Name: ')
        cus_id = input('Customer ID: ')
        cus_address = input('Address: ')
        cus_product = input('Product: ')

        #Get valid amount
        while True:
            cus_amount = input('Amount: ')
            is_valid, float_amount = validate_amount(cus_amount)
            if is_valid:
                break
            else:
                print('Invalid amount. Please enter a valid numeric amount.')

        cus_pay_method = get_valid_input('Payment Method(Credit Card/Debit Card/Cash/Online Payment): ', validate_payment_method)
        cus_order_deliver = get_valid_input('Order Delivered (Yes/No): ', validate_order_delivered)
        cus_remark = input('Remark: ')

        #Split the date and time
        cus_date, cus_time = cus_date_n_time.split()

        #Add the data to the worksheet
        next_row = worksheet.max_row + 1
        worksheet.cell(row=next_row, column=1, value=cus_date)
        worksheet.cell(row=next_row, column=2, value=cus_time)
        worksheet.cell(row=next_row, column=3, value=order_id)
        worksheet.cell(row=next_row, column=4, value=cus_id)
        worksheet.cell(row=next_row, column=5, value=cus_name)
        worksheet.cell(row=next_row, column=6, value=cus_product)
        worksheet.cell(row=next_row, column=7, value=float_amount)
        worksheet.cell(row=next_row, column=8, value=cus_pay_method)
        worksheet.cell(row=next_row, column=9, value=cus_order_deliver)
        worksheet.cell(row=next_row, column=10, value=cus_remark)

        #Save the workbook
        workbook.save('Bakery.xlsx')
        print(' ')
        print('=' * 60)
        print('Data has been saved to the Excel file with the order number:', order_id)
        print('Returning to main menu...')
        print(' ')

        pass

    #Option 6 - Print customer bill
    elif option == '006':
        print('-'*60)
        print('Option 006 - Print bill statement for customer')
        print('-'*60)

        #Function to search for a customer by customer ID or name
        def search_customer(worksheet, masterList_worksheet, search_key):
            customer_data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                customer_id = row[3]
                if customer_id == search_key:
                    row = list(row)  #Convert the tuple to a list
                    product_code = row[5]
                    product_name = get_product_name(masterList_worksheet, product_code)
                    row.append(product_name)  #Add the product name to the list
                    row = tuple(row)  #Convert the list back to a tuple
                    customer_data.append(row)
            return customer_data

        #Function to retrieve the product name from the "MasterList" worksheet based on the product code
        def get_product_name(masterList_worksheet, product_code):
            for row in masterList_worksheet.iter_rows(min_row=2, values_only=True):
                code = row[0]
                if code == product_code:
                    return row[2]

        #Function to print a customer's bill statement
        def print_bill_statement(customer_data):
            if not customer_data:
                print("Customer not found.")
            else:
                for row in customer_data:
                    print('               ~~SUNFLOWER BAKERY SDN BHD~~')
                    print('                  Lot 101, Jalan Kampar,')
                    print('                      31900, Kampar.')
                    print(' ')
                    print('-' * 60)
                    print('                 CUSTOMER BILL STATEMENT')
                    print('-' * 60)
                    print(f"Order No:                   {row[2]}")
                    print(f"Customer ID:                {row[3]}")
                    print(f"Customer Name:              {row[4]}")
                    print(f"Date:                       {row[0]}")
                    print(f"Time:                       {row[1]}")
                    print(' ')
                    
                    #Print the order details
                    print('-' * 60)
                    print('                     DELIVERY ORDER')
                    print('-' * 60)
                    print(f"Product ID:                 {row[5]}")
                    print(f"Product Name:               {row[10]}")
                    print(f"Amount:                     {row[6]}")
                    print(f"Payment Method:             {row[7]}")
                    print(f"Order Delivered:            {row[8]}")
                    print(f"Remarks:                    {row[9]}")
                    print(' ')
                    
                    #Get the current date and time
                    from datetime import datetime
                    current_datetime = datetime.now()
                    formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

                    #Print the footer of the bill
                    print('-' * 60)
                    print('             Printed on: ', formatted_datetime)
                    print('               ~~SUNFLOWER BAKERY SYSTEM~~')
                    print(' ')
                    print('=' * 60)
                    print('Customer bill statement for customer has been printed')
                    print('Returning to main menu...')
                    print(' ')

        #Create or open the Excel workbook
        workbook = openpyxl.load_workbook('Bakery.xlsx')

        #Select the worksheet
        worksheet = workbook['Customer_Detail']

        #Select the "MasterList" worksheet
        masterList_worksheet = workbook['MasterList']

        #Prompt the user to enter a customer ID or name for searching
        search_key = input("Enter Customer ID to search: ")
        print(' ')

        #Search for the customer
        customer_data = search_customer(worksheet, masterList_worksheet, search_key)

        #Print the customer's bill statement
        print_bill_statement(customer_data)

        pass

    #Option 7 - Generate Report for each months
    elif option == '007':
        print('-'*60)
        print('Option 007 - Generate Report for each months')
        print('-'*60)

        from collections import defaultdict
        from datetime import datetime

        #Function to filter data for a specific month and year
        def filter_data_by_month(worksheet, year, month):
            filtered_data = []
            is_header = True  #To track if it's the header row
            for row in worksheet.iter_rows(min_row=4, values_only=True):
                date_string = row[0]
                if is_header:
                    is_header = False  #Skip the header row
                    continue
                if date_string is not None:  #Check if date_string is not None
                    try:
                        date_obj = datetime.strptime(date_string, '%Y-%m-%d')
                        if date_obj.year == year and date_obj.month == month:
                            filtered_data.append(row)
                    except ValueError:
                        pass  #Ignore rows with invalid date formats
            return filtered_data

        #Function to generate a monthly report
        def generate_monthly_report(filtered_data):
            monthly_report = defaultdict(lambda: {'total_sales': 0, 'num_orders': 0})
            for row in filtered_data:
                month_year = row[0][:7]  #Extract YYYY-MM part from the date
                amount = row[6]  #Assuming the amount is in the 7th column
                monthly_report[month_year]['total_sales'] += amount
                monthly_report[month_year]['num_orders'] += 1
            return monthly_report

        #Open the Excel workbook
        workbook = openpyxl.load_workbook('Bakery.xlsx')

        #Select the "Customer_Detail" worksheet
        worksheet = workbook['Customer_Detail']

        #Prompt the user to enter the year and month for the report
        year = int(input("Enter the year (e.g., 2023): "))
        month = int(input("Enter the month (01-12): "))
        print(' ')

        #Filter data for the specified month and year
        filtered_data = filter_data_by_month(worksheet, year, month)

        #Generate the monthly report
        monthly_report = generate_monthly_report(filtered_data)

        #Print the monthly report
        print('-' * 60)
        print(f"Monthly Report for {year}-{month:02d}")
        print('-' * 60)
        for month_year, data in monthly_report.items():
            print(f"Month-Year: {month_year}")
            print(f"Total Sales: ${data['total_sales']:.2f}")
            print(f"Number of Orders: {data['num_orders']}")
            print(' ')
            print('=' * 60)
            print('Customer bill statement for customer has been printed')
            print('Returning to main menu...')
            print(' ')

        pass

    #Option 8 - Exit system
    elif option == '008':     
        print('-'*60)
        print('        Thank you for using Sunflower Bakery System!')
        print('-'*60)

        break

    else:
        print('Invalid option. Please select a valid option.')
